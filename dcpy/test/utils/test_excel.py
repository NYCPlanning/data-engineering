from pathlib import Path

import openpyxl  # type: ignore
import pytest

from dcpy.utils import excel

RESOURCES_DIR = Path(__file__).parent / "resources"

BASE_TAB = "my_base_tab"
BASE_KEYS = ["geo_id", "geo_type", "year"]
BASE_HEADER_ROW = 3

MODIFICATIONS_TAB = "my_modifications"
MODIFICATIONS_KEYS = ["geo_id", "geo_typeZZZ", "year"]
MODIFICATIONS_HEADER_ROW = 4

UNUSED_CORRECTIONS_VARIABLE = "unused_corrs"


@pytest.fixture
def default_mod_kwargs(tmp_path):
    xlsx = openpyxl.load_workbook(RESOURCES_DIR / "excel_with_updates.xlsx")
    out_path = tmp_path / "modified.xlsx"
    return {
        "base_wb": xlsx,
        "base_tab": BASE_TAB,
        "base_keys": BASE_KEYS,
        "base_header_row": BASE_HEADER_ROW,
        "modifications_wb": xlsx,
        "modifications_tab": MODIFICATIONS_TAB,
        "modifications_keys": MODIFICATIONS_KEYS,
        "modifications_header_row": MODIFICATIONS_HEADER_ROW,
        "out_path": out_path,
    }


@pytest.fixture
def default_insert_csv_into_sheet_kwargs(tmp_path):
    return {
        "csv_path": RESOURCES_DIR / "data_for_excel.csv",
        "input_excel_path": RESOURCES_DIR / "template_for_csvs.xlsx",
        "output_excel_path": tmp_path / "template_for_csvs_modified.xlsx",
        "sheet_name": "First sheet",
        "row_offset": 6,
    }


def test_applying_modifications(default_mod_kwargs: dict):
    excel.apply_cross_file_modifications(**default_mod_kwargs)
    modified_xlsx = openpyxl.load_workbook(default_mod_kwargs["out_path"])

    modified_vals = excel._get_table_values_by_keys(
        wb=modified_xlsx, tab_name=BASE_TAB, keys=BASE_KEYS, header_row=BASE_HEADER_ROW
    )
    for _, v in modified_vals.items():
        # Every value that SHOULD be changed in the base tab will have a starting value of
        # "change me"
        assert "change me" not in v.values()


def test_applying_modifications__missing_keys(default_mod_kwargs: dict):
    nonsense_key = "asdfasdfasdfasd"
    with pytest.raises(Exception) as e:
        default_mod_kwargs["modifications_keys"] = [nonsense_key]
        excel.apply_cross_file_modifications(**default_mod_kwargs)

    assert str(e.value).startswith(excel.MISSING_KEYS_ERROR)
    assert nonsense_key in str(e.value), (
        "The error message should mention the missing key."
    )


def test_applying_modifications__duplicate_keys(default_mod_kwargs: dict):
    with pytest.raises(Exception) as e:
        default_mod_kwargs["modifications_keys"] = ["year"]
        excel.apply_cross_file_modifications(**default_mod_kwargs)

    assert str(e.value).startswith(excel.DUPLICATE_KEYS_ERROR)


def test_applying_modifications__missing_destination_columns(default_mod_kwargs: dict):
    with pytest.raises(Exception) as e:
        default_mod_kwargs["allow_missing_destination_columns"] = False
        excel.apply_cross_file_modifications(**default_mod_kwargs)
        assert False

    assert str(e.value).startswith(excel.MISSING_DESTINATION_CORR_COL_ERROR)
    assert UNUSED_CORRECTIONS_VARIABLE in str(e.value)


def test_insert_csv_into_sheet(default_insert_csv_into_sheet_kwargs: dict):
    excel.insert_csv_into_sheet(**default_insert_csv_into_sheet_kwargs)

    modified_excel = openpyxl.load_workbook(
        default_insert_csv_into_sheet_kwargs["output_excel_path"]
    )

    assert (
        modified_excel[default_insert_csv_into_sheet_kwargs["sheet_name"]][
            default_insert_csv_into_sheet_kwargs["row_offset"] + 1
        ][0].value
        == "a"
    )
    assert (
        modified_excel[default_insert_csv_into_sheet_kwargs["sheet_name"]][
            default_insert_csv_into_sheet_kwargs["row_offset"] + 2
        ][0].value
        == "data_1"
    )


def test_insert_csv_into_sheet__missing_sheet_name(
    default_insert_csv_into_sheet_kwargs: dict,
):
    default_insert_csv_into_sheet_kwargs["sheet_name"] = None
    excel.insert_csv_into_sheet(**default_insert_csv_into_sheet_kwargs)

    modified_excel = openpyxl.load_workbook(
        default_insert_csv_into_sheet_kwargs["output_excel_path"]
    )

    assert (
        modified_excel.active[default_insert_csv_into_sheet_kwargs["row_offset"] + 1][
            0
        ].value
        == "a"
    )
    assert (
        modified_excel.active[default_insert_csv_into_sheet_kwargs["row_offset"] + 2][
            0
        ].value
        == "data_1"
    )


def test_insert_csv_into_sheet__missing_row_offset(
    default_insert_csv_into_sheet_kwargs: dict,
):
    default_insert_csv_into_sheet_kwargs["sheet_name"] = "Second sheet"
    default_insert_csv_into_sheet_kwargs.pop("row_offset")
    excel.insert_csv_into_sheet(**default_insert_csv_into_sheet_kwargs)

    modified_excel = openpyxl.load_workbook(
        default_insert_csv_into_sheet_kwargs["output_excel_path"]
    )

    assert (
        modified_excel[default_insert_csv_into_sheet_kwargs["sheet_name"]][1][0].value
        == "a"
    )
    assert (
        modified_excel[default_insert_csv_into_sheet_kwargs["sheet_name"]][4][3].value
        == "stuff"
    )
