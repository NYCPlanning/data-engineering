"""
Generate district XLSX files from EDDE packager outputs.

This module takes the resolved JSON files from the packager and generates
XLSX files for each geography (boroughs and community districts).
"""

import json
from pathlib import Path
from typing import Any

import openpyxl
from config import get_edde_paths
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dcpy.utils.logging import logger

# Get build data paths dynamically
OLD_EDDE_PATH, NEW_BUILD_PATH = get_edde_paths()
PACKAGE_DIR = NEW_BUILD_PATH.parent / "package"
RESOLVED_DIR = PACKAGE_DIR / "resolved_pages_and_tables" / "districts"
OUTPUT_DIR = PACKAGE_DIR / "district_xlsx"

# Create output directory
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# Category mappings
CATEGORY_INFO = {
    "demo": {
        "sheet_name": "Demographic Conditions",
        "subcategories": ["tot", "anh", "bnh", "wnh", "hsp"],
    },
    "econ": {
        "sheet_name": "Household Economic Security",
        "subcategories": ["tot", "anh", "bnh", "hsp", "wnh"],
    },
    "hsaq": {
        "sheet_name": "Housing Afrdblty, Qlty, Scrty",
        "full_name": "Housing Affordability, Quality, and Security",
        "subcategories": ["tot", "anh", "bnh", "wnh", "hsp"],
    },
    "hopd": {
        "sheet_name": "Housing Production",
        "subcategories": ["tot"],
    },
    "qlao": {
        "sheet_name": "QoL and Access to Opportunity",
        "full_name": "Quality of Life and Access to Opportunity",
        "subcategories": ["tot", "anh", "bnh", "wnh", "hsp"],
    },
}

# Geography name mappings
GEOGRAPHY_NAMES = {
    "borough_1": "Bronx",
    "borough_2": "Brooklyn",
    "borough_3": "Manhattan",
    "borough_4": "Queens",
    "borough_5": "SI",  # Staten Island
}

# Subcategory name suffixes for table titles
SUBCATEGORY_SUFFIXES = {
    "tot": "Total Population",
    "anh": "Asian Non-Hispanic",
    "bnh": "Black Non-Hispanic",
    "hsp": "Hispanic",
    "wnh": "White Non-Hispanic",
}

# Mapping from subcategory code to table ID suffix
SUBCATEGORY_TO_TABLE_SUFFIX = {
    "tot": "T",
    "anh": "A",
    "bnh": "B",
    "hsp": "H",
    "wnh": "W",
}


def create_styled_workbook() -> openpyxl.Workbook:
    """Create a new workbook with basic setup."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def add_sheet_title(
    ws: openpyxl.worksheet.worksheet.Worksheet, title: str, geography: str
):
    """Add the main title at the top of a sheet."""
    ws.cell(1, 1).value = f"{title}\n{geography}"
    ws.cell(1, 1).font = Font(name="Calibri", size=24)
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40


def add_table_to_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    table: dict[str, Any],
    start_row: int,
    subcategory: str = "tot",
    add_suffix: bool = True,
) -> int:
    """
    Add a single table to the worksheet.

    Returns the next available row number.
    """
    current_row = start_row

    # Table title with subcategory suffix
    table_id = table["id"]
    table_title_base = table["title"]

    # Add subcategory suffix only if requested and not the race/ethnicity breakdown table
    if add_suffix and table_id != "1.01.T":
        suffix = SUBCATEGORY_SUFFIXES.get(subcategory, "")
        table_title = f"Table {table_id}: {table_title_base} - {suffix}"
    else:
        table_title = f"Table {table_id}: {table_title_base}"

    ws.cell(current_row, 1).value = table_title
    ws.cell(current_row, 1).font = Font(name="Calibri", size=11, bold=True)
    current_row += 1

    # Build header structure from vintages
    vintages = table.get("vintages", [])
    if not vintages:
        return current_row + 2  # Skip empty table

    # Determine number of columns needed
    # For each vintage, we have columns based on headers
    col_offset = 2  # Start at column B (column 1 is for row labels)

    # First header row: vintage labels
    for vintage in vintages:
        if vintage.get("isChange"):
            label = f"Change, {vintages[-2]['label'].split(', ')[-1]} to {vintages[-1]['label'].split(', ')[-1]}"
        else:
            label = vintage.get("label", "")
            # Convert to title case for consistency with original XLSX
            label = label.title()

        # Get number of columns for this vintage from headers
        num_cols = len(vintage.get("headers", [[]])[0]) if vintage.get("headers") else 2

        ws.cell(current_row, col_offset).value = label
        if num_cols > 1:
            ws.merge_cells(
                start_row=current_row,
                start_column=col_offset,
                end_row=current_row,
                end_column=col_offset + num_cols - 1,
            )

        # Style header
        for i in range(num_cols):
            cell = ws.cell(current_row, col_offset + i)
            cell.fill = PatternFill(
                start_color="2C7A7B", end_color="2C7A7B", fill_type="solid"
            )
            cell.font = Font(name="Calibri", size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        col_offset += num_cols

    current_row += 1

    # Second header row: measure types (Number, Percent, etc.)
    col_offset = 2
    for vintage in vintages:
        headers = vintage.get("headers", [[]])
        if headers and len(headers) > 0:
            for header_item in headers[0]:
                ws.cell(current_row, col_offset).value = header_item["label"].title()
                ws.cell(current_row, col_offset).font = Font(name="Calibri", size=11)
                ws.cell(current_row, col_offset).alignment = Alignment(
                    horizontal="center"
                )
                col_offset += header_item.get("colspan", 1)

    current_row += 1

    # Data rows
    first_vintage = vintages[0]
    rows = first_vintage.get("rows", [])

    for row_data in rows:
        # Row label
        ws.cell(current_row, 1).value = row_data.get("label", "")
        ws.cell(current_row, 1).font = Font(name="Calibri", size=11)

        # Data cells across all vintages
        col_offset = 2
        for vintage in vintages:
            # Find matching row in this vintage
            matching_row = next(
                (
                    r
                    for r in vintage.get("rows", [])
                    if r.get("label") == row_data.get("label")
                ),
                None,
            )

            if matching_row:
                cells = matching_row.get("cells") or []
                for cell_data in cells:
                    value = cell_data.get("value")
                    cell = ws.cell(current_row, col_offset)
                    cell.value = value
                    cell.font = Font(name="Calibri", size=11)

                    # Apply number formatting
                    measure = cell_data.get("measure")
                    if measure == "COUNT" and value is not None:
                        cell.number_format = "#,##0"
                    elif measure == "PERCENT" and value is not None:
                        # Check if it's already a percentage or needs conversion
                        cell.number_format = "0.0"

                    col_offset += 1
            else:
                # No data for this vintage, skip columns
                num_cols = (
                    len(vintage.get("headers", [[]])[0])
                    if vintage.get("headers")
                    else 2
                )
                col_offset += num_cols

        current_row += 1

    # Add spacing after table
    return current_row + 2


def generate_xlsx_for_geography(
    resolved_dir: Path,
    geography_key: str,
    output_path: Path,
):
    """
    Generate an XLSX file for a specific geography.

    Args:
        resolved_dir: Path to the resolved_pages_and_tables/districts directory
        geography_key: Geography identifier (e.g., "borough_5")
        output_path: Where to save the output XLSX file
    """
    geography_name = GEOGRAPHY_NAMES.get(geography_key, geography_key)

    wb = create_styled_workbook()

    for category, info in CATEGORY_INFO.items():
        json_file = resolved_dir / f"{geography_key}_{category}.json"

        if not json_file.exists():
            logger.warning(f"{json_file} not found, skipping {category}")
            continue

        with open(json_file) as f:
            data = json.load(f)

        # Create sheet
        sheet_name = info["sheet_name"]
        full_name = info.get("full_name", sheet_name)
        ws = wb.create_sheet(title=sheet_name)

        # Add title
        add_sheet_title(ws, full_name, geography_name)

        # Set column widths
        ws.column_dimensions["A"].width = 30.33
        for col_num in range(2, 23):  # Columns B-V
            ws.column_dimensions[get_column_letter(col_num)].width = 13.0

        # Hide gridlines
        ws.sheet_view.showGridLines = False

        # Add tables from all subcategories
        # The order matters - need to interleave by table ID prefix
        current_row = 19  # Start tables at row 19 (matching original format)

        # Determine if this category has multiple subcategories with different data
        # (i.e., if it's broken down by race/ethnicity)
        has_race_ethnicity_breakdown = len(info["subcategories"]) > 1

        # Collect all tables from all subcategories
        all_tables = []
        for subcat_key in info["subcategories"]:
            subcat_tables = data.get(subcat_key, [])
            for table in subcat_tables:
                table_id = table.get("id")
                # Extract the suffix (last character after the last dot)
                expected_suffix = SUBCATEGORY_TO_TABLE_SUFFIX.get(subcat_key)

                # Only include this table if its ID suffix matches the subcategory
                # e.g., table "1.02.A" should only come from "anh" subcategory
                if table_id and "." in table_id:
                    actual_suffix = table_id.split(".")[-1]
                    if actual_suffix == expected_suffix:
                        all_tables.append((table_id, table, subcat_key))

        # Sort tables by ID with custom ordering: T first, then A, B, H, W
        def table_sort_key(item):
            table_id = item[0]
            if not table_id or "." not in table_id:
                return (0, "")

            # Split into prefix (e.g., "1.02") and suffix (e.g., "T")
            parts = table_id.split(".")
            prefix = ".".join(parts[:-1])
            suffix = parts[-1]

            # Define sort order for suffixes: T first, then A, B, H, W
            suffix_order = {"T": 0, "A": 1, "B": 2, "H": 3, "W": 4}
            suffix_rank = suffix_order.get(suffix, 99)

            return (prefix, suffix_rank)

        all_tables.sort(key=table_sort_key)

        # Render all tables
        for table_id, table, subcat_key in all_tables:
            current_row = add_table_to_sheet(
                ws,
                table,
                current_row,
                subcategory=subcat_key,
                add_suffix=has_race_ethnicity_breakdown,
            )

    # Save workbook
    wb.save(output_path)
    logger.info(f"Generated XLSX: {output_path}")


def main():
    """Generate XLSX files for all geographies."""
    logger.info(f"Resolved data directory: {RESOLVED_DIR}")
    logger.info(f"Output directory: {OUTPUT_DIR}")

    # Get all geography JSON files
    if not RESOLVED_DIR.exists():
        logger.error(f"Resolved directory not found: {RESOLVED_DIR}")
        return

    # Find all unique geographies from the JSON files
    geographies = set()
    for json_file in RESOLVED_DIR.glob("*.json"):
        # Extract geography key from filename (e.g., "borough_5" from "borough_5_demo.json")
        parts = json_file.stem.split("_")
        if len(parts) >= 2:
            # Handle both "borough_5" and multi-part like "cd_101"
            geo_key = "_".join(parts[:-1])
            geographies.add(geo_key)

    logger.info(f"Found {len(geographies)} geographies to process")

    # Generate XLSX for each geography
    for geo_key in sorted(geographies):
        # Determine output filename
        if geo_key.startswith("borough_"):
            borough_num = geo_key.split("_")[1]
            display_name = GEOGRAPHY_NAMES.get(geo_key, f"Borough{borough_num}")
            output_filename = f"{display_name}.xlsx"
        else:
            # For community districts or other geographies
            output_filename = f"{geo_key}.xlsx"

        output_path = OUTPUT_DIR / output_filename

        logger.info(f"Generating {output_filename} for {geo_key}")
        generate_xlsx_for_geography(
            resolved_dir=RESOLVED_DIR,
            geography_key=geo_key,
            output_path=output_path,
        )

    logger.info(f"Successfully generated {len(geographies)} XLSX files")


if __name__ == "__main__":
    main()
