from copy import copy
import openpyxl  # type: ignore
from openpyxl.styles import Border, Side, Alignment  # type: ignore
from openpyxl.cell.text import InlineFont  # type: ignore
from openpyxl.cell.rich_text import TextBlock, CellRichText  # type: ignore
from openpyxl.drawing.image import Image  # type: ignore
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder  # type: ignore
from openpyxl.utils import get_column_letter
from pathlib import Path
import typer

from dcpy.lifecycle import product_metadata
from dcpy.models.product.metadata import OrgMetadata
from dcpy.models.design import elements as de
from dcpy.utils.logging import logger

from . import RESOURCES_PATH
from . import abstract_doc

# TODO: Move template to Product Metadata Repo. Rename to be non-OTI specific
DEFAULT_TEMPLATE_PATH = RESOURCES_PATH / "oti_data_dictionary_template.xlsx"
EXCEL_DATA_DICT_METADATA_FILE_TYPE = "oti_data_dictionary"
DEFAULT_FONT = "Arial"


def _set_default_style(cell, *, is_rightmost=True, is_topmost=False, is_last_row=False):
    border_side_thin = Side(border_style="thin", color="000000")
    border_side_medium = Side(border_style="medium", color="000000")

    cell.alignment = Alignment(wrapText=True, vertical="center")
    cell.border = Border(
        top=border_side_medium if is_topmost else border_side_thin,
        left=border_side_thin,
        right=border_side_medium if is_rightmost else border_side_thin,
        bottom=border_side_medium if is_last_row else border_side_thin,
    )


def _format_row_slice(row_slice, is_first_row=False, is_last_row=False):
    """Format a row slice with OTI's table formatting."""
    *left_cells, rightmost_cell = row_slice
    [
        _set_default_style(
            c,
            is_last_row=is_last_row,
            is_topmost=is_first_row,
            is_rightmost=False,
        )
        for c in left_cells
    ]
    _set_default_style(
        rightmost_cell,
        is_topmost=is_first_row,
        is_last_row=is_last_row,
        is_rightmost=True,
    )


def _abstract_style_to_xlsx(c: de.CellStyle):
    return {
        k: v
        for k, v in {
            "rFont": c.font.name or DEFAULT_FONT,
            "color": c.font.rgb,
            "b": c.font.bold,
            "sz": c.font.size,
            "i": c.font.italic,
        }.items()
        if v
    }


def _to_human_readable_val(v) -> str:
    if type(v) is bool:
        return "Yes" if v else "No"
    elif not v:
        return ""
    else:
        return str(v)


def generate_table_sheet(
    xlsx_wb: openpyxl.Workbook,
    table: de.Table,
    *,
    tab_name: str,
    tab_index: int = -1,
    table_row_start_index=1,
):
    """Adds a worksheet + table to and XLSX file."""
    new_sheet = xlsx_wb.create_sheet(title=tab_name, index=tab_index)
    new_sheet.sheet_view.showGridLines = False

    new_sheet.insert_rows(table_row_start_index, len(table.rows))
    new_sheet.insert_cols(1, table.total_cols() - 1)
    new_sheet_rows = [r for r in new_sheet.rows]

    # Set Column Widths when specified
    dim_holder = DimensionHolder(worksheet=new_sheet)  # type: ignore
    for idx, col in enumerate(range(new_sheet.min_column, new_sheet.max_column + 1)):
        col_dim = ColumnDimension(new_sheet, min=col, max=col)

        maybe_width = (
            table.column_widths[idx]
            if table.column_widths and len(table.column_widths) > idx
            else None
        )
        if maybe_width:
            col_dim.width = maybe_width

        dim_holder[get_column_letter(col)] = col_dim
    new_sheet.column_dimensions = dim_holder

    for r_idx, r in enumerate(table.rows):
        row = new_sheet_rows[r_idx]
        if r.merge_cells:
            # for merged cells, just format the top-leftmost cell
            if not r.skip_default_styling:
                _format_row_slice(
                    row[0:1],
                    is_first_row=r.is_top_row,
                    is_last_row=r_idx == len(table.rows) - 1,
                )
            new_sheet.merge_cells(
                start_row=r_idx + 1,
                end_row=r_idx + 1,
                start_column=1,
                end_column=table.total_cols(),
            )
        else:
            if not r.skip_default_styling:
                _format_row_slice(row, is_last_row=r_idx == len(table.rows) - 1)

        if r.height:
            new_sheet.row_dimensions[r_idx + 1].height = r.height
        for c_idx, c in enumerate(r.cells):
            if type(c.value) is list and c.value and type(c.value[0]) is de.Cell:
                # Inline Cells
                cell = CellRichText(
                    [
                        TextBlock(
                            InlineFont(**_abstract_style_to_xlsx(ic.style)),
                            _to_human_readable_val(ic.value),
                        )
                        for ic in c.value
                    ]
                )
                row[c_idx].value = cell
            elif type(c.value) is de.Image:
                cell = row[c_idx]

                # TODO: should probably use: https://openpyxl.readthedocs.io/en/3.1/api/openpyxl.utils.units.html
                PIXELS_PER_INCH = 96.0
                img = Image(c.value.path)
                img.height = int(1.01 * PIXELS_PER_INCH)
                img.width = int(6.15 * PIXELS_PER_INCH)
                new_sheet.add_image(img, cell.coordinate)  # type: ignore
            else:
                row[c_idx].value = CellRichText(
                    TextBlock(
                        InlineFont(**_abstract_style_to_xlsx(c.style)),
                        _to_human_readable_val(c.value),
                    )
                )
            if c.style.text_alignment_vertical or c.style.text_alignment_horizontal:
                alignment = copy(row[c_idx].alignment)
                if c.style.text_alignment_vertical:
                    alignment.vertical = c.style.text_alignment_vertical
                if c.style.text_alignment_horizontal:
                    alignment.horizontal = c.style.text_alignment_horizontal
                row[c_idx].alignment = alignment


def write_xlsx(
    *,
    org_md: OrgMetadata,
    product: str,
    dataset: str | None = None,
    output_path: Path | None = None,
    template_path_override: Path | None = None,
    artifact_name: str | None = None,
):
    """Adds Metadata Tables to an Excel Template.

    For a given product.dataset, this will add a worksheet + table for each component specified
    in the artifacts.yml in the metadata repo.
    """

    artifacts = org_md.get_packaging_artifacts()
    dataset = dataset or product

    xlsx_artifacts = [a for a in artifacts if a.type == "xlsx"]
    artifact = None
    if artifact_name:
        matched = [a for a in xlsx_artifacts if a.name == artifact_name]
        if len(matched) != 1:
            raise Exception(f"Expected exactly one artifact named {matched}")
        artifact = matched[0]
    else:
        if len(xlsx_artifacts) == 1:
            artifact = xlsx_artifacts[0]
        else:
            raise Exception("An artifact name must be specified")

    tables = abstract_doc.generate_abstract_artifact(
        product=product, dataset=dataset, org_metadata=org_md, artifact=artifact
    )
    xlsx_wb = openpyxl.load_workbook(
        filename=template_path_override or DEFAULT_TEMPLATE_PATH
    )

    tab_and_tables = zip(artifact.components, tables)
    for tab, table in tab_and_tables:
        generate_table_sheet(
            xlsx_wb, table=table, tab_index=tab.index, tab_name=tab.name
        )

    if "delete_me" in xlsx_wb.sheetnames:
        # A saved worksheet requires at least one visible tab, so we left one called `delete_me`
        del xlsx_wb["delete_me"]
    out_path = output_path or Path("./data_dictionary.xlsx")
    logger.info(f"Saving OTI XLSX to {out_path}")
    xlsx_wb.save(out_path)


app = typer.Typer()


@app.command("generate_xlsx")
def _write_xlsx_cli(
    product: str,
    dataset: str,
    version: str,
    artifact_name: Path = typer.Option(
        None,
        "--artifact-name",
        "-n",
        help="Name of the xlsx artifact to generate",
    ),
    output_path: Path = typer.Option(
        Path("data_dictionary.xlsx"),
        "--output-path",
        "-o",
        help="Output Path. Defaults to ./data_dictionary.xlsx",
    ),
    template_path_override: Path = typer.Option(
        None,
        "--template-path",
        "-t",
        help="(Override) Template Path",
    ),
):
    write_xlsx(
        product=product,
        dataset=dataset,
        org_md=product_metadata.load(version=version),
        output_path=output_path,
        template_path_override=template_path_override,
    )
