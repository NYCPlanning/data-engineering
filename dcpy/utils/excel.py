import typer
import openpyxl  # type: ignore
import csv
from pathlib import Path

from dcpy.utils.logging import logger

MISSING_KEYS_ERROR = "Supplied keys must be found in the target table"
MISSING_DESTINATION_CORR_COL_ERROR = "Destination is missing column for correction"
DUPLICATE_KEYS_ERROR = "Found duplicate keys"


def _get_table_values_by_keys(
    wb: openpyxl.Workbook, tab_name: str, keys: list[str], *, header_row=1
) -> dict[tuple, dict]:
    """Transform a table into a dictionary of (key_1, key_2, ...)->values."""
    tab = wb[tab_name]
    headers = [str(c.value) for c in tab[header_row]]
    assert set(keys).issubset(set(headers)), (
        f"{MISSING_KEYS_ERROR}. Diff: {set(keys) - set(headers)}"
    )

    unpivoted = {}
    duplicate_keys = set()
    for i, row in enumerate(tab.iter_rows()):
        if i < header_row:  # skip the headers
            continue
        row_dict = dict(zip(headers, [c.value for c in row]))
        key = tuple([str(row_dict[k]) for k in keys])
        if key in unpivoted:
            duplicate_keys.add(key)
        unpivoted[key] = {k: v for k, v in row_dict.items() if k not in keys}
    if duplicate_keys:
        raise Exception(f"{DUPLICATE_KEYS_ERROR}. Keys: {duplicate_keys}")
    return unpivoted


def _update_table_cells_by_keys(
    wb: openpyxl.Workbook,
    tab_name: str,
    keys: list[str],
    update_values,
    *,
    header_row=1,
    allow_missing_destination_columns=True,
):
    """"""
    tab = wb[tab_name]
    headers = [c.value for c in tab[header_row]]
    assert set(keys).issubset(set(headers))

    sample_key = next(iter(update_values))
    sample_value_keys = update_values[sample_key].keys()
    missing_keys = set(sample_value_keys) - set(headers)
    logger.warning(f"{MISSING_DESTINATION_CORR_COL_ERROR}. Missing: {missing_keys}")
    if not allow_missing_destination_columns:
        assert not missing_keys, (
            f"{MISSING_DESTINATION_CORR_COL_ERROR}. Missing: {missing_keys}"
        )

    updated_rows = 0
    updated_cells = 0
    logger.info("updating rows and cells")
    for i, row in enumerate(tab.iter_rows()):
        if i < header_row:  # headers
            continue
        row_dict = dict(zip(headers, [c.value for c in row]))
        key = tuple([str(row_dict[k]) for k in keys])

        if key in update_values:
            update_vals = update_values[key]
            for k, v in update_vals.items():
                if v is not None and k in headers:
                    value_idx = headers.index(k)
                    row[value_idx].value = v
                    updated_cells += 1
            updated_rows += 1
    print(f"updated rows: {updated_rows}, updated cells: {updated_cells}")


def apply_cross_file_modifications(
    *,
    base_wb: openpyxl.Workbook,
    base_tab: str,
    base_keys: list[str],
    modifications_wb: openpyxl.Workbook,
    modifications_tab: str,
    modifications_keys: list[str],
    out_path: Path,
    base_header_row=1,
    modifications_header_row=1,
    allow_missing_destination_columns=True,
):
    """Edit the `base` workbook tab using values from the `modifications` file.

    Values from the `modifications` will be accumulated using the `modifications_keys`,
    then matched and inserted into the `base` tab, joining on the `base_keys.`
    """
    logger.info(f"Unpivoting vars in base wb, tab={modifications_tab}")
    unpivoted_vars = _get_table_values_by_keys(
        modifications_wb,
        modifications_tab,
        modifications_keys,
        header_row=modifications_header_row,
    )
    logger.info(
        f"Unpivoted. Found {len(unpivoted_vars.keys())} that will be used to update ."
    )

    _update_table_cells_by_keys(
        wb=base_wb,
        tab_name=base_tab,
        keys=base_keys,
        update_values=unpivoted_vars,
        allow_missing_destination_columns=allow_missing_destination_columns,
        header_row=base_header_row,
    )
    base_wb.save(out_path)


def csv_into_excel(
    csv_path: Path,
    input_excel_path: Path,
    output_excel_path: Path,
    sheet_name: str | None = None,
    row_ofset: int | None = None,
):
    """
    Use a CSV file to populate cells in an Excel file.
    """
    row_ofset = row_ofset or 0
    workbook = openpyxl.load_workbook(filename=input_excel_path)
    if sheet_name:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    with open(csv_path, "r") as csvfile:
        csv_reader = csv.reader(csvfile)
        for row_index, row in enumerate(csv_reader, start=1):
            for col_index, cell_value in enumerate(row, start=1):
                sheet.cell(
                    row=row_ofset + row_index,
                    column=col_index,
                    value=cell_value,
                )

    workbook.save(output_excel_path)


app = typer.Typer()


@app.command("apply_modifications")
def apply_mods(
    base_wb_path: str = typer.Option(..., help="Path to base workbook"),
    base_tab: str = typer.Option(..., help="Base tab name"),
    base_keys: list[str] = typer.Option([], "-b", "--base", help="Base key columns"),
    mods_wb_path: str = typer.Option(..., help="Path to modifications workbook"),
    mods_tab: str = typer.Option(..., help="Modifications tab name"),
    mods_keys: list[str] = typer.Option([], "-m", help="Modification key columns"),
    out_path: Path = typer.Option(..., help="Output path"),
):
    """Apply modifications from one Excel workbook to another"""
    logger.info(f"loading base workbook at {base_wb_path}")
    base_wb = openpyxl.load_workbook(base_wb_path)
    logger.info(f"loading modifications workbook at {mods_wb_path}")
    mods_wb = openpyxl.load_workbook(mods_wb_path)
    logger.info(base_keys)
    apply_cross_file_modifications(
        base_wb=base_wb,
        base_tab=base_tab,
        base_keys=base_keys,
        modifications_wb=mods_wb,
        modifications_tab=mods_tab,
        modifications_keys=mods_keys,
        out_path=out_path,
    )
